import os
import socket
import pymysql
import datetime
import numpy as np
import pandas as pd
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont
from openpyxl.styles import Font
from PyQt5 import QtWidgets, QtGui
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QWidget, QMessageBox, QInputDialog


FONT_12 = QFont('Kaiti TC', 15)
VERSION = 'project manage system v1.0.0'
NAME = VERSION + ' --- by elepikachu'
TITLE = 'manage mode'


# -------------------------------------------------------------
# 函数名： conn_database
# 功能： 连接数据库
# -------------------------------------------------------------
def conn_database():
    db = pymysql.connect(host="localhost", port=3306, user="root", password="123456", database="elepikachu")
    return db


class Manage_UI(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    # -------------------------------------------------------------
    # 函数名： initUI
    # 功能： UI初始化
    # -------------------------------------------------------------
    def initUI(self):
        self.setGeometry(300, 300, 600, 400)
        self.setWindowTitle(TITLE)
        self.setWindowIcon(QtGui.QIcon('jng.ico'))
        self.setToolTip(NAME)
        self.img_route = 'background2.jpeg'
        if os.path.exists(self.img_route):
            self.use_palette(img_route=self.img_route)

        self.helpButton = QtWidgets.QPushButton(self)
        self.helpButton.setGeometry(0, 0, 100, 25)
        self.helpButton.setText("Help")
        self.helpButton.setToolTip("app help")
        self.helpButton.clicked.connect(self.help_window)

        self.logButton = QtWidgets.QPushButton(self)
        self.logButton.setGeometry(100, 0, 100, 25)
        self.logButton.setText("Log")
        self.logButton.setToolTip("Open log file")
        self.logButton.clicked.connect(self.log_window)

        self.exitButton = QtWidgets.QPushButton(self)
        self.exitButton.setGeometry(500, 0, 100, 25)
        self.exitButton.setText("Exit")
        self.exitButton.setToolTip("close app")
        self.exitButton.clicked.connect(self.close)

        self.titleLabel = QtWidgets.QLabel(self)
        self.titleLabel.setText(VERSION + ' ' + TITLE)
        self.titleLabel.setFont(FONT_12)
        self.titleLabel.setGeometry(75, 40, 450, 20)
        self.titleLabel.setStyleSheet("Color:green")

        self.dataTable = QtWidgets.QTableWidget(self)
        row = self.tableInit()
        self.dataTable.setAlternatingRowColors(True)
        self.dataTable.setStyleSheet("background:rgb(255,255,255,0.8)")
        self.dataTable.setGeometry(50, 80, 500, 200)
        self.dataTable.cellChanged.connect(self.cellChange)

        self.outputBox = QtWidgets.QTextBrowser(self)
        self.outputBox.setStyleSheet("background:rgb(255,255,255,0.6)")
        self.outputBox.setGeometry(300, 300, 250, 75)

        self.updateButton = QtWidgets.QPushButton(self)
        self.updateButton.setGeometry(40, 310, 75, 20)
        self.updateButton.setText("save line")
        self.updateButton.setToolTip("update the line you selected")
        self.updateButton.clicked.connect(self.lineUpdate)

        self.deleteButton = QtWidgets.QPushButton(self)
        self.deleteButton.setGeometry(130, 310, 75, 20)
        self.deleteButton.setText("delete line")
        self.deleteButton.setToolTip("delete the line you selected")
        self.deleteButton.clicked.connect(self.lineDelete)

        self.refreshButton = QtWidgets.QPushButton(self)
        self.refreshButton.setGeometry(220, 310, 75, 20)
        self.refreshButton.setText("refresh")
        self.refreshButton.setToolTip("refresh the table status")
        self.refreshButton.clicked.connect(lambda: self.refresh_table('all', ''))

        self.nameFilterBox = QtWidgets.QComboBox(self)
        self.nameFilterBox.setGeometry(40, 365, 75, 20)
        self.groupFilterBox = QtWidgets.QComboBox(self)
        self.groupFilterBox.setGeometry(130, 365, 165, 20)
        self.setFilterBox(row)

        self.filterButton = QtWidgets.QPushButton(self)
        self.filterButton.setGeometry(40, 340, 75, 20)
        self.filterButton.setText("use filter")
        self.filterButton.setToolTip("filter data according to name/group, you can only set one filter")
        self.filterButton.clicked.connect(self.use_filter)

        self.filterButton = QtWidgets.QPushButton(self)
        self.filterButton.setGeometry(130, 340, 75, 20)
        self.filterButton.setText("delete all")
        self.filterButton.setToolTip("delete everything in the chart")
        self.filterButton.clicked.connect(self.delete_all)

        self.printButton = QtWidgets.QPushButton(self)
        self.printButton.setGeometry(220, 340, 75, 20)
        self.printButton.setText("print chart")
        self.printButton.setToolTip("print the chart according to the filter")
        self.printButton.clicked.connect(self.create_excel)

    # -------------------------------------------------------------
    # 函数名： tableInit
    # 功能： 表格初始化
    # -------------------------------------------------------------
    def tableInit(self):
        db = conn_database()
        cursor = db.cursor()
        cursor.execute("select * from buyitem_item")
        all = cursor.fetchall()
        row = len(all)
        vol = len(all[0])
        col_lst = [tup[0] for tup in cursor.description]
        self.dataTable.setColumnCount(vol)
        self.dataTable.setRowCount(row)
        self.dataTable.setHorizontalHeaderLabels(col_lst)
        self.dataTable.verticalHeader().setVisible(False)
        for i in range(row):
            for j in range(vol):
                tmp = all[i][j]
                data = QtWidgets.QTableWidgetItem(str(tmp))
                self.dataTable.setItem(i, j, data)
        for i in range(row):
            self.dataTable.item(i, 0).setFlags(Qt.ItemIsEnabled)
        return row

    # -------------------------------------------------------------
    # 函数名： cellChange
    # 功能： 修改表格后提醒保存
    # -------------------------------------------------------------
    def cellChange(self):
        self.outputBox.setText('chart changed, please click save line after edit')

    # -------------------------------------------------------------
    # 函数名： setFilterBox
    # 功能： 过滤项初始化
    # -------------------------------------------------------------
    def setFilterBox(self, row):
        name_list = ['all']
        self.nameFilterBox.clear()
        for i in range(row):
            name = self.dataTable.item(i, 5).text()
            if name not in name_list:
                name_list.append(name)
        self.nameFilterBox.addItems(name_list)
        group_list = ['all']
        self.groupFilterBox.clear()
        for i in range(row):
            group = self.dataTable.item(i, 10).text()
            if group not in group_list:
                group_list.append(group)
        self.groupFilterBox.addItems(group_list)

    # -------------------------------------------------------------
    # 函数名： lineUpdate
    # 功能： 单行更新
    # -------------------------------------------------------------
    def lineUpdate(self):
        db = conn_database()
        cursor = db.cursor()
        reply = QMessageBox.question(self, 'Message', 'Are you sure to change it ?', QMessageBox.Yes | QMessageBox.No,
                                     QMessageBox.No)
        if reply == QMessageBox.Yes:
            row_2 = self.dataTable.currentRow()
            col = 14
            upd_d = []
            for i in range(col):
                upd_d.append(self.dataTable.item(row_2, i).text())

            # 在数据库修改数据
            cmd = f"update buyitem_item set good='{upd_d[1]}', brand='{upd_d[2]}', unit='{upd_d[3]}', quantity='{upd_d[4]}', name='{upd_d[5]}', phone='{upd_d[6]}', num='{upd_d[7]}', info='{upd_d[8]}', detail='{upd_d[9]}', `group`='{upd_d[10]}', date='{upd_d[11]}', finish='{upd_d[12]}', classif='{upd_d[13]}' WHERE id = '{upd_d[0]}';"
            try:
                cursor.execute(cmd)
                db.commit()
            except Exception as e:
                print(e)
            cursor.execute('select * from buyitem_itemlog order by id desc limit 1;')
            maxIdx2 = cursor.fetchone()
            idx2 = maxIdx2[0] + 1
            ip = socket.gethostbyname(socket.gethostname())
            date = datetime.datetime.today()
            cmd = f"insert into buyitem_itemlog value({idx2}, '{ip}', '{date}', 'update', '{upd_d[5]}-{upd_d[5]}-{upd_d[12]}');"
            try:
                cursor.execute(cmd)
                db.commit()
            except Exception as e:
                print(e)
            self.outputBox.setText('update success, please refresh')

    # -------------------------------------------------------------
    # 函数名： lineDelete
    # 功能： 单行删除
    # -------------------------------------------------------------
    def lineDelete(self):
        db = conn_database()
        cursor = db.cursor()
        reply = QMessageBox.question(self, 'Message', 'Are you sure to delete it ?', QMessageBox.Yes | QMessageBox.No,
                                     QMessageBox.No)
        if reply == QMessageBox.Yes:
            row_2 = self.dataTable.currentRow()
            del_d = self.dataTable.item(row_2, 0).text()
            name = self.dataTable.item(row_2, 5).text()
            good = self.dataTable.item(row_2, 1).text()

            # 在数据库删除数据
            try:
                cursor.execute(f"DELETE FROM buyitem_item WHERE id = '{del_d}';")
                db.commit()
            except Exception as e:
                print(e)
            cursor.execute('select * from buyitem_itemlog order by id desc limit 1;')
            maxIdx2 = cursor.fetchone()
            idx2 = maxIdx2[0] + 1
            ip = socket.gethostbyname(socket.gethostname())
            date = datetime.datetime.today()
            cmd = f"insert into buyitem_itemlog value({idx2}, '{ip}', '{date}', 'delete', '{name}-{good}');"
            try:
                cursor.execute(cmd)
                db.commit()
            except Exception as e:
                print(e)
            self.outputBox.setText('delete success, please refresh')

    # -------------------------------------------------------------
    # 函数名： delete_all
    # 功能： 删除全部
    # -------------------------------------------------------------
    def delete_all(self):
        db = conn_database()
        cursor = db.cursor()
        reply = QInputDialog.getInt(self, "qtmanage", "Danger Operation! please input manager passcode before next step!")
        if reply == 666666:
            try:
                cursor.execute('truncate table;')
                db.commit()
            except Exception as e:
                print(e)
            cursor.execute('select * from buyitem_itemlog order by id desc limit 1;')
            maxIdx2 = cursor.fetchone()
            idx2 = maxIdx2[0] + 1
            ip = socket.gethostbyname(socket.gethostname())
            date = datetime.datetime.today()
            cmd = f"insert into buyitem_itemlog value({idx2}, '{ip}', '{date}', 'delete', 'all');"
            try:
                cursor.execute(cmd)
                db.commit()
            except Exception as e:
                print(e)
            self.outputBox.setText('delete success, please refresh')

    # -------------------------------------------------------------
    # 函数名： refresh_table
    # 功能： 刷新表格，重新从数据库取数据
    # -------------------------------------------------------------
    def refresh_table(self, classif, name):
        db = conn_database()
        cursor = db.cursor()
        if classif == 'all':
            cursor.execute("select * from buyitem_item")
        elif classif == 'name':
            cursor.execute(f"select * from buyitem_item where name='{name}'")
        else:
            cursor.execute(f"select * from buyitem_item where `group`='{name}'")
        all = cursor.fetchall()
        row = len(all)
        vol = len(all[0])
        self.dataTable.setRowCount(row)
        for i in range(row):
            for j in range(vol):
                tmp = all[i][j]
                data = QtWidgets.QTableWidgetItem(str(tmp))
                self.dataTable.setItem(i, j, data)
        for i in range(row):
            self.dataTable.item(i, 0).setFlags(Qt.ItemIsEnabled)
        if classif == 'all':
            self.setFilterBox(row)
            self.outputBox.setText('refresh success')

    # -------------------------------------------------------------
    # 函数名： use_filter
    # 功能： 应用过滤
    # -------------------------------------------------------------
    def use_filter(self):
        name = self.nameFilterBox.currentText()
        group = self.groupFilterBox.currentText()
        if name == 'all' and group == 'all':
            self.refresh_table('all', '')
            self.outputBox.setText('clear filter success')
        elif name == 'all':
            self.refresh_table('group', group)
            self.outputBox.setText('use filter success')
        elif group == 'all':
            self.refresh_table('name', name)
            self.outputBox.setText('use filter success')
        else:
            self.outputBox.setText('You can not set both name and group filter')

    # -------------------------------------------------------------
    # 函数名： create_excel
    # 功能： 打印表格
    # -------------------------------------------------------------
    def create_excel(self):
        reply = QMessageBox.question(self, 'Message', 'You could set filter before print chart',
                                     QMessageBox.Yes | QMessageBox.No,
                                     QMessageBox.No)
        if reply == QMessageBox.No:
            return
        db = conn_database()
        cursor = db.cursor()
        name = self.nameFilterBox.currentText()
        group = self.groupFilterBox.currentText()
        mon = (datetime.datetime.today()).strftime("%m")
        mon = str(int(mon) + 1)
        if name == 'all' and group == 'all':
            cursor.execute("select * from buyitem_item")
            fname = f'股份月度计划-%s月.xlsx' % (mon)
            office = False
        elif group == 'all':
            cursor.execute(f"select * from buyitem_item where name='{name}'")
            fname = f'股份月度计划-%s月-%s.xlsx' % (mon, name)
            office = True
        elif name == 'all':
            cursor.execute(f"select * from buyitem_item where `group`='{group}'")
            fname = f'股份月度计划-%s月-%s.xlsx' % (mon, group)
            office = True
        else:
            self.outputBox.setText('You can not set both name and group filter')
            return
        all = cursor.fetchall()
        rawdata = pd.DataFrame(all)
        rawdata.columns = ['序号', '商品名称', '品牌型号', '单位', '数量', '姓名', '电话', '课题编号', '采购说明',
                           '备注', '11', '12', '13', '14']

        writer = pd.ExcelWriter(fname, engine='openpyxl')
        writer = self.print_excel(writer, '股份-办公用品采购', '办公用品', rawdata, office)
        writer = self.print_excel(writer, '股份-设备耗材采购', '设备耗材', rawdata, office)
        writer = self.print_excel(writer, '股份-办公家具', '办公家具', rawdata, office)
        writer = self.print_excel(writer, '股份-五金杂品采购', '五金杂品', rawdata, office)
        writer = self.print_excel(writer, '股份-劳动防护用品', '劳动防护', rawdata, office)
        writer = self.print_excel(writer, '股份-实验耗材及小型设备', '实验耗材及小型设备', rawdata, office)
        writer.save()
        self.outputBox.setText('print success, file name is %s' % fname)

    # -------------------------------------------------------------
    # 函数名： print_excel
    # 功能： 按类别把数据输出到excel
    # -------------------------------------------------------------
    def print_excel(self, writer, word, wd, rawdata, of):
        data = rawdata[rawdata['14'] == wd]
        if not data.empty:
            data = data.drop(['11', '12', '13', '14'], axis=1)
            data.style.apply(excel_style, axis=0).to_excel(writer, sheet_name=word, index=False, startrow=3)

            column_wid = (data.columns.to_series().apply(lambda x: len(x.encode('gbk'))).values)
            max_wid = (data.astype(str).applymap(lambda x: len(x.encode('gbk'))).agg(max).values)
            wids = np.max([column_wid, max_wid], axis=0)
            worksheet = writer.sheets[word]
            for i, wid in enumerate(wids, 1):
                worksheet.column_dimensions[get_column_letter(i)].width = wid + 2
        else:
            try:
                data = data.drop(['11', '12', '13', '14'], axis=1)
            except:
                pass

            data.to_excel(writer, sheet_name=word, index=False, startrow=3)

        worksheet = writer.sheets[word]
        mon = (datetime.datetime.today()).strftime("%m")
        mon = str(int(mon) + 1)
        worksheet.cell(row=1, column=1).value = '            中国石油勘探开发研究院采购计划表(' + mon + '月)'
        worksheet['A1'].font = Font(size=20, bold=True)
        worksheet.cell(row=2, column=1).value = '表单号：NKZ0400-01'
        worksheet['A2'].font = Font(size=12, bold=True)
        worksheet.cell(row=2, column=6).value = '类别：' + wd
        worksheet['F2'].font = Font(size=12, bold=True)
        if of:
            worksheet.cell(row=3, column=1).value = '申请单位：' + rawdata['11'][0]
        else:
            worksheet.cell(row=3, column=1).value = '申请单位：'
        worksheet['A3'].font = Font(size=12, bold=True)
        worksheet.cell(row=3, column=6).value = '填报日期：' + datetime.datetime.today().strftime("%Y-%m-%d")
        worksheet['F3'].font = Font(size=12, bold=True)
        return writer

    # -------------------------------------------------------------
    # 函数名： use_palette
    # 功能： 使用背景
    # -------------------------------------------------------------
    def use_palette(self, img_route):
        window_pale = QtGui.QPalette()
        pix = QtGui.QPixmap(img_route)
        pix = pix.scaled(600, 400)
        window_pale.setBrush(QtGui.QPalette.Background, QtGui.QBrush(pix))
        self.setPalette(window_pale)

    # -------------------------------------------------------------
    # 函数名： help_window
    # 功能： 打开帮助窗口
    # -------------------------------------------------------------
    def help_window(self):
        self.window5 = Help()
        self.window5.show_info()
        self.window5.show()

    # -------------------------------------------------------------
    # 函数名： help_window
    # 功能： 打开日志
    # -------------------------------------------------------------
    def log_window(self):
        self.window6 = Log()
        self.window6.get_log()
        self.window6.show()

    # -------------------------------------------------------------
    # 函数名： closeEvent
    # 功能： 关闭时询问
    # -------------------------------------------------------------
    def closeEvent(self, event):
        reply = QMessageBox.question(self, 'Warning', 'Are you sure to quit', QMessageBox.Yes, QMessageBox.No)
        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()


class Help(QWidget):
    def __init__(self):
        super().__init__()

        self.setGeometry(300, 300, 400, 600)
        self.setWindowTitle("Help")
        self.setWindowIcon(QtGui.QIcon('jng.png'))

        self.titleLabel = QtWidgets.QLabel(self)
        self.titleLabel.setText(TITLE + ' - Help')
        self.titleLabel.setFont(FONT_12)
        self.titleLabel.setGeometry(75, 50, 300, 20)
        self.titleLabel.setStyleSheet("Color:green")

        self.outBox = QtWidgets.QTextBrowser(self)
        self.outBox.setGeometry(50, 100, 300, 400)
        self.outBox.setStyleSheet("background:rgb(255,255,255,0.6)")
        self.show()

    # -------------------------------------------------------------
    # 函数名： show_info
    # 功能： 读取帮助信息
    # -------------------------------------------------------------
    def show_info(self):
        with open(r'Help.txt', 'r', encoding='utf-8') as f:
            text = f.read()
        self.outBox.setText(text)


class Log(QWidget):
    def __init__(self):
        super().__init__()

        self.setGeometry(300, 300, 400, 600)
        self.setWindowTitle("Log")
        self.setWindowIcon(QtGui.QIcon('jng.png'))

        self.titleLabel = QtWidgets.QLabel(self)
        self.titleLabel.setText(TITLE + ' - Log')
        self.titleLabel.setFont(FONT_12)
        self.titleLabel.setGeometry(75, 50, 300, 20)
        self.titleLabel.setStyleSheet("Color:green")

        self.outBox = QtWidgets.QTextBrowser(self)
        self.outBox.setGeometry(50, 100, 300, 400)
        self.outBox.setStyleSheet("background:rgb(255,255,255,0.6)")
        self.show()


    def get_log(self):
        db = conn_database()
        cursor = db.cursor()
        cursor.execute("select * from buyitem_itemlog")
        text = str(cursor.fetchall())
        self.outBox.setText(text)


# -------------------------------------------------------------
# 函数名： excel_style
# 功能： 添加excel的格式
# -------------------------------------------------------------
def excel_style(x):
    return ['text-align:center' for x in x]
